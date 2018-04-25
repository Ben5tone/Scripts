'''
Created on 20/02/2018

@author: JCHAV106
'''
import xlrd
from tkinter import messagebox, Toplevel
from openpyxl import load_workbook
import xml.etree.ElementTree as ET
import os
import sys
import ctypes.wintypes

CSIDL_PERSONAL = 5       # My Documents
SHGFP_TYPE_CURRENT = 0   # Get current, not default value

buf= ctypes.create_unicode_buffer(ctypes.wintypes.MAX_PATH)
ctypes.windll.shell32.SHGetFolderPathW(None, CSIDL_PERSONAL, None, SHGFP_TYPE_CURRENT, buf)

class Extract(object):
    
    
    def __init__(self,):
        ""

    def exl2xml(self,xmlf,exlf,xmlfo):
    
        tree = ET.parse(xmlf)
        root = tree.getroot()
    
        #Open the workbook to read in
        x1_workbook = xlrd.open_workbook(exlf)
    
        #Obtain spreadsheet names of the current workbook
        sheet_names = x1_workbook.sheet_names()

        print('Sheet Names',sheet_names)
        
        #Call the first sheet of the book
        x1_sheet = x1_workbook.sheet_by_index(0)
    
        #List variables to put the names of the bushing names contained in the xml and excel file
        xml_bushings = []
        exl_bushings = []
        #Iterate over inline.xml file to obtain all the bushing names 
        for child in root.iter("NVHC_PROPERTY"):
            for child2 in child.iter("NAME"):
                xml_bushings.insert(0, child2.text)
                #print(xml_bushings)

                cont = 0
                bushings_dict = {}

            for row_idx in range(0, x1_sheet.nrows):    # Iterate through rows
                for col_idx in range(0, x1_sheet.ncols):  # Iterate through columns
                    cell_obj = x1_sheet.cell(row_idx, col_idx).value  # Get cell object by row, col 
                    if cell_obj == "Bushing":
                        for col_idx in range(col_idx, x1_sheet.ncols):    # Iterate through rows
                            for row_idx in range(row_idx, x1_sheet.nrows): # Iterate through columns
                                cell_obj2 = x1_sheet.cell(row_idx, col_idx).value
                                for i in range(xml_bushings.__len__()):
                                    if cell_obj2 == xml_bushings[i]:
                                        r = row_idx 
                                        #print(cell_obj2, xml_bushings[i], i )
                                        for col_idx2 in range(col_idx+1, 10):    # Iterate through rows
                                            for row_idx in range(col_idx+1, 18):  # Iterate through columns
                                                cell_obj3 = x1_sheet.cell(r, col_idx2).value
                                                #print(cell_obj3,cell_obj2, xml_bushings[i], a + 1)
                                                exl_bushings.insert(exl_bushings.__len__(), cell_obj3)
                                                cont = cont + 1
                                                if cont == 18:
                                                    cont = 0
                                                    bushings_dict[str(cell_obj2)] = exl_bushings
                                                    exl_bushings = [] 
                                                    break
                                    
                                                break

        print(bushings_dict)

    #Iterate trough all the NVHC PROPERTY tags contained in the file
        for child3 in root.iter("NVHC_PROPERTY"):
            #Iterate trough all the NAME tags that are child of each NVHC PROPERTY tag
            for child4 in child3.iter("NAME"):
                #Iterate trough items of the bushing_dict to obtain their keys and values
                for key, value in bushings_dict.items():
                    #Iterate to compare if the NAME obtained from the Bushing matches with one of the keys of the bushings_dict
                    if child4.text == key:
                        print('?????????????????????????')
                        print(key)
                        #Accesign to each element of the list that is inside the found key
                        rt1 = bushings_dict[child4.text][0]
                        rt2 = bushings_dict[child4.text][1] 
                        rt3 = bushings_dict[child4.text][2] 
                        rt4 = bushings_dict[child4.text][3] 
                        rt5 = bushings_dict[child4.text][4] 
                        rt6 = bushings_dict[child4.text][5]
                        rt7 = bushings_dict[child4.text][6]
                        rt8 = bushings_dict[child4.text][7] 
                        rt9 = bushings_dict[child4.text][8] 
                        rt10 = bushings_dict[child4.text][9] 
                        rt11 = bushings_dict[child4.text][10] 
                        rt12 = bushings_dict[child4.text][11]
                        rt13 = bushings_dict[child4.text][12]
                        rt14 = bushings_dict[child4.text][13] 
                        rt15 = bushings_dict[child4.text][14] 
                        rt16 = bushings_dict[child4.text][15] 
                        rt17 = bushings_dict[child4.text][16] 
                        rt18 = bushings_dict[child4.text][17]
                        #Iterate trough all the PARAMETERS tags that are child of each NVHC PROPERTY tag  
                        for child4 in child3.iter("PARAMETERS"):
                            #Iterates trough the content of the PARAMMETERS tag to find the busjing values(stiffnes(k) and ge) and replace each 
                            #one of them with the corresponding value obtained from the excel file
                            for child5 in child4:
                                #Compare if the child tag of PARAMETERS matches with one of the values that gonna be replaced
                                if child5.tag == "K_VALUES":
                                    child5.set("b1", str(rt1))
                                    child5.set("b2", str(rt2))
                                    child5.set("b3", str(rt3))
                                    child5.set("b4", str(rt4))
                                    child5.set("b5", str(rt5))
                                    child5.set("b6", str(rt6))
                                    #Write the new data to the same xml file or create a new with diferent name 
                                    tree.write(xmlfo)  
                                    #print(child5.attrib)
                            
                                elif child5.tag == "B_VALUES":
                                    child5.set("ge1", str(rt7))
                                    child5.set("ge2", str(rt8))
                                    child5.set("ge3", str(rt9))
                                    child5.set("ge4", str(rt10))
                                    child5.set("ge5", str(rt11))
                                    child5.set("ge6", str(rt12))
                                    #Write the new data to the same xml file or create a new with diferent name  
                                    tree.write(xmlfo)  
                                    #print(child5.attrib)
                                
                                elif child5.tag == "GE_VALUES":
                                    child5.set("ge1", str(rt13))
                                    child5.set("ge2", str(rt14))
                                    child5.set("ge3", str(rt15))
                                    child5.set("ge4", str(rt16))
                                    child5.set("ge5", str(rt17))
                                    child5.set("ge6", str(rt18))
                                    #Write the new data to the same xml file or create a new with diferent name  
                                    tree.write(xmlfo)  
                                    #print(child5.attrib)
        messagebox.showinfo('Process','Done')
    def xml2exl(self,xmlf,exlfo):
        
        if getattr(sys, 'frozen', False):
            # If the application is run as a bundle, the pyInstaller bootloader
            # extends the sys module by a flag frozen=True and sets the app 
            # path into variable _MEIPASS'.
            dirname = os.path.dirname(sys.executable)
        else:
            dirname = os.path.dirname(__file__)
            
        template = os.path.join(dirname,'Extract_bushing_rates_tool.xlsm')
        #save = os.path.join(dirname,'Extract_bushing_rates_tool.xlsx')
        
        tree = ET.parse(xmlf)
        root = tree.getroot()
    
        #Open the workbook to read in
        #x1_workbook = xlrd.open_workbook(exlf)
    
        #Obtain spreadsheet names of the current workbook
        #sheet_names = x1_workbook.sheet_names()

        #print('Sheet Names',sheet_names)
        
        #Call the first sheet of the book
        #x1_sheet = x1_workbook.sheet_by_index(5)
       
        #Open the workbook to read in
        y1_workbook = load_workbook(template)
        wb = y1_workbook['CM1']
    
        #List variables to put the names of the bushing names contained in the xml and excel file
        xml_bushings = {}
   
        save_num = 0
        dic = {}                         
        num = 0
        varn = ""
        varn2 = []
        varsbs1 = []
        varsbs2 = {}
        bushings_list = []
        prop = []
        cont_varn = 0
        numero = 0
        bushingsss = {}
        bushsave = {}
        for control_group in root:
            if control_group.tag == "NVHC_CONTROL_GROUP":
                for subsys in control_group:
                    if subsys.tag == "NAME":
                        sub_key = subsys.text
                        dic[sub_key]= {}
                        num += 1
                        if num == 1:
                            cont = 0
                            cont_varn = 0
                            if varn2 == []:
                                dic[sub_key][""]={}
                                dic[sub_key][""][""]=bushings_list
                            else:
                                for varnames in varn2:
                                    dic[sub_key][varnames] = {}
                                    dic[sub_key][varnames] = varsbs2
                                    varsbs2 = {}
                        else:
                            dic[sub_key][""]={}
                            dic[sub_key][""][""]=bushings_list
                        
                        bushings_list = []
                        varn2 = []
                    if subsys.tag == "CONTENTS":
                        for contents in subsys:              
                            if contents.tag == "NVHC_CONTROL_GROUP":    
                                for varname in contents:
                                        if varname.tag == "NAME":  
                                            varn = varname.text
                                            varn2.insert(varn2.__len__(),varn)
                                            if num == 0:
                                                print("")
                                            else:
                                                varsbs1 = []
                                            varsbs1.insert(varsbs1.__len__(),varsbs2)
                                            num = 0
                                        if varname.tag == "CONTENTS":
                                            for control_group2 in varname:
                                                if control_group2.tag == "NVHC_CONTROL_GROUP":
                                                    for varsubs in control_group2:
                                                        if varsubs.tag == "NAME":
                                                            print(varsubs.text)
                                                            varsbs2[varsubs.text]= prop
                                                            prop = []
                                                            print(varsbs2)  
                                                        if varsubs.tag == "CONTENTS":                                                
                                                            for varsubs2 in varsubs:
                                                                if varsubs2.tag == "NVHC_CONNECTOR":
                                                                    for varsubs3 in varsubs2:
                                                                        if varsubs3.tag == "REPRESENTATION":
                                                                            for varsubs4 in varsubs3:
                                                                                for varsubs5 in varsubs4:
                                                                                    if varsubs5.tag == "PROPERTY":
                                                                                        prop.insert(prop.__len__(),varsubs5.text)
                            if contents.tag == "NVHC_PROPERTY":
                                for bushings in contents:
                                    if bushings.tag == "NAME":
                                        bushings_list.insert(0,bushings.text) 
        print(dic)
        
        for key, value in dic.items():
            numero += 1
            print(key)
            wb.cell(row = 3 + numero, column = 1, value = key)
            for key2, value2 in value.items():
                numero += 1
                print(key2)
                print(value2.items())
                wb.cell(row = 3 + numero, column = 2, value = key2)
                for key3, value3 in value2.items():
                    numero +=1
                    print("--------------")
                    print(key3)
                    print(value3)
                    wb.cell(row = 3 + numero, column = 3, value = key3)
                    for key4 in value3:
                        numero +=1
                        print(key4)
                        wb.cell(row = 3 + numero, column = 4, value = key4)
        
        if exlfo != None:
            y1_workbook.save(exlfo)
        else:
            #save_path = 'Extract_bushing_rates_tool' + '(' + str(save_num) + ')' + '.xlsx'
            exists = os.path.exists(buf.value + '/Extract_bushings_rates_tool' + '(' + str(save_num) + ')' + '.xlsx')
            while exists == True:
                save_num += 1
                #save_path = 'Extract_bushing_rates_tool' + '(' + str(save_num) + ')' + '.xlsx'
                exists = os.path.exists(buf.value + '/Extract_bushings_rates_tool' + '(' + str(save_num) + ')' + '.xlsx')
            save = os.path.join(dirname,buf.value + '/Extract_bushings_rates_tool' + '(' + str(save_num) + ')' + '.xlsx')
            y1_workbook.save(save)

        if exlfo != None:
            x1_workbook = xlrd.open_workbook(exlfo)
        else:

            x1_workbook = xlrd.open_workbook(save)

        x1_sheet = x1_workbook.sheet_by_index(0)

        for control_group2 in root:
            if control_group2.tag == "NVHC_CONTROL_GROUP":
                for subsys2 in control_group2:
                    if subsys2.tag == "CONTENTS":
                        for contents2 in subsys2:  
                            if contents2.tag == "NVHC_PROPERTY":
                                for bushings2 in contents2:
                                    if bushings2.tag == "NAME":
                                        bushname = bushings2.text 
                                        bushingsss[bushname] = {}
                                    if bushings2.tag =="PARAMETERS" :
                                        for values2 in bushings2:
                                            'bushsave[values2.tag] = values2.attrib'
                                            bushingsss[bushname][values2.tag] = {}
                                            bushingsss[bushname][values2.tag] = values2.attrib
        print(bushingsss)
        for b_k, b_val in bushingsss.items():
            for row_idx in range(0, x1_sheet.nrows):
                if x1_sheet.cell(row_idx,3).value == b_k:
                    for col_idx in range(5,x1_sheet.ncols+1):
                        if col_idx == 5:
                            wb.cell(row = row_idx+1, column = col_idx, value = float(bushingsss[b_k]['K_VALUES']['k1']))
                        if col_idx == 6:
                            wb.cell(row = row_idx+1, column = col_idx, value = float(bushingsss[b_k]['K_VALUES']['k2']))
                        if col_idx == 7:
                            wb.cell(row = row_idx+1, column = col_idx, value = float(bushingsss[b_k]['K_VALUES']['k3']))
                        if col_idx == 8:
                            wb.cell(row = row_idx+1, column = col_idx, value = float(bushingsss[b_k]['K_VALUES']['k4']))
                        if col_idx == 9:
                            wb.cell(row = row_idx+1, column = col_idx, value = float(bushingsss[b_k]['K_VALUES']['k5']))
                        if col_idx == 10:
                            wb.cell(row = row_idx+1, column = col_idx, value = float(bushingsss[b_k]['K_VALUES']['k6']))
                        if col_idx == 11:
                            wb.cell(row = row_idx+1, column = col_idx, value = float(bushingsss[b_k]['B_VALUES']['b1']))
                        if col_idx == 12:
                            wb.cell(row = row_idx+1, column = col_idx, value = float(bushingsss[b_k]['B_VALUES']['b2']))
                        if col_idx == 13:
                            wb.cell(row = row_idx+1, column = col_idx, value = float(bushingsss[b_k]['B_VALUES']['b3']))
                        if col_idx == 14:
                            wb.cell(row = row_idx+1, column = col_idx, value = float(bushingsss[b_k]['B_VALUES']['b4']))
                        if col_idx == 15:
                            wb.cell(row = row_idx+1, column = col_idx, value = float(bushingsss[b_k]['B_VALUES']['b5']))
                        if col_idx == 16:
                            wb.cell(row = row_idx+1, column = col_idx, value = float(bushingsss[b_k]['B_VALUES']['b6']))
                        if col_idx == 17:
                            wb.cell(row = row_idx+1, column = col_idx, value = float(bushingsss[b_k]['GE_VALUES']['ge1']))
                        if col_idx == 18:
                            wb.cell(row = row_idx+1, column = col_idx, value = float(bushingsss[b_k]['GE_VALUES']['ge2']))
                        if col_idx == 19:
                            wb.cell(row = row_idx+1, column = col_idx, value = float(bushingsss[b_k]['GE_VALUES']['ge3']))
                        if col_idx == 20:
                            wb.cell(row = row_idx+1, column = col_idx, value = float(bushingsss[b_k]['GE_VALUES']['ge4']))
                        if col_idx == 21:
                            wb.cell(row = row_idx+1, column = col_idx, value = float(bushingsss[b_k]['GE_VALUES']['ge5']))
                        if col_idx == 22:
                            wb.cell(row = row_idx+1, column = col_idx, value = float(bushingsss[b_k]['GE_VALUES']['ge6']))
        if exlfo != None:
            y1_workbook.save(exlfo)
            messagebox.showinfo('Process','Done')
        else:
            
            #save_path = 'Extract_bushing_rates_tool' + '(' + str(save_num) + ')' + '.xlsx'
            exists = os.path.exists(buf.value + '/Extract_bushings_rates_tool' + '(' + str(save_num) + ')' + '.xlsx')
            while exists == True:
                save_num += 1
                #save_path = 'Extract_bushing_rates_tool' + '(' + str(save_num) + ')' + '.xlsx'
                exists = os.path.exists(buf.value + '/Extract_bushings_rates_tool' + '(' + str(save_num) + ')' + '.xlsx')
            save = os.path.join(dirname,buf.value + '/Extract_bushings_rates_tool' + '(' + str(save_num) + ')' + '.xlsx')
            y1_workbook.save(save)
            messagebox.showinfo('Process','Done')