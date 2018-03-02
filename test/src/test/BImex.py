'''
Created on 20/02/2018

@author: JCHAV106
'''
import xlrd
from tkinter import messagebox, Toplevel
from openpyxl import load_workbook
import xml.etree.ElementTree as ET

class BImex(object):
    
    
    def __init__(self,):
        ""

    def exl2xml(self,xmlf,exlf,xmlfo,s_num):
    
        tree = ET.parse(xmlf)
        root = tree.getroot()
    
        #Open the workbook to read in
        x1_workbook = xlrd.open_workbook(exlf)
    
        #Obtain spreadsheet names of the current workbook
        sheet_names = x1_workbook.sheet_names()

        print('Sheet Names',sheet_names)
        
        #Call the first sheet of the book
        x1_sheet = x1_workbook.sheet_by_index(s_num)
    
        #List variables to put the names of the bushing names contained in the xml and excel file
        xml_bushings = []
        exl_bushings = []
        #Iterate over inline.xml file to obtain all the bushing names 
        for child in root.iter("NVHC_PROPERTY"):
            for child2 in child.iter("NAME"):
                xml_bushings.insert(0, child2.text)
                #print(xml_bushings)

                cont = 0
                bushings_dict = {}

            for row_idx in range(0, x1_sheet.nrows):    # Iterate through rows
                for col_idx in range(0, x1_sheet.ncols):  # Iterate through columns
                    cell_obj = x1_sheet.cell(row_idx, col_idx).value  # Get cell object by row, col 
                    if cell_obj == "Bushing":
                        for col_idx in range(col_idx, x1_sheet.ncols):    # Iterate through rows
                            for row_idx in range(row_idx, x1_sheet.nrows): # Iterate through columns
                                cell_obj2 = x1_sheet.cell(row_idx, col_idx).value
                                for i in range(xml_bushings.__len__()):
                                    if cell_obj2 == xml_bushings[i]:
                                        r = row_idx 
                                        #print(cell_obj2, xml_bushings[i], i )
                                        for col_idx2 in range(col_idx+1, 10):    # Iterate through rows
                                            for row_idx in range(col_idx+1, 6):  # Iterate through columns
                                                cell_obj3 = x1_sheet.cell(r, col_idx2).value
                                                #print(cell_obj3,cell_obj2, xml_bushings[i], a + 1)
                                                exl_bushings.insert(0, cell_obj3)
                                                cont = cont + 1
                                                if cont == 6:
                                                    cont = 0
                                                    bushings_dict[str(cell_obj2)] = exl_bushings
                                                    exl_bushings = [] 
                                                    break
                                    
                                                break

        print(bushings_dict)

    #Iterate trough all the NVHC PROPERTY tags contained in the file
        for child3 in root.iter("NVHC_PROPERTY"):
            #Iterate trough all the NAME tags that are child of each NVHC PROPERTY tag
            for child4 in child3.iter("NAME"):
                #Iterate trough items of the bushing_dict to obtain their keys and values
                for key, value in bushings_dict.items():
                    #Iterate to compare if the NAME obtained from the Bushing matches with one of the keys of the bushings_dict
                    if child4.text == key:
                        print(key)
                        #Accesign to each element of the list that is inside the found key
                        rt1 = bushings_dict[child4.text][0]
                        rt2 = bushings_dict[child4.text][1] 
                        rt3 = bushings_dict[child4.text][2] 
                        rt4 = bushings_dict[child4.text][3] 
                        rt5 = bushings_dict[child4.text][4] 
                        rt6 = bushings_dict[child4.text][5]
                        #Iterate trough all the PARAMETERS tags that are child of each NVHC PROPERTY tag  
                        for child4 in child3.iter("PARAMETERS"):
                            #Iterates trough the content of the PARAMMETERS tag to find the busjing values(stiffnes(k) and ge) and replace each 
                            #one of them with the corresponding value obtained from the excel file
                            for child5 in child4:
                                #Compare if the child tag of PARAMETERS matches with one of the values that gonna be replaced
                                if child5.tag == "B_VALUES":
                                    child5.set("b1", str(0))
                                    child5.set("b2", str(0))
                                    child5.set("b3", str(0))
                                    child5.set("b4", str(0))
                                    child5.set("b5", str(0))
                                    child5.set("b6", str(0))
                                    #Write the new data to the same xml file or create a new with diferent name 
                                    tree.write(xmlfo)  
                                    #print(child5.attrib)
                            
                                elif child5.tag == "K_VALUES":
                                    child5.set("k1", str(rt6))
                                    child5.set("k2", str(rt5))
                                    child5.set("k3", str(rt4))
                                    child5.set("k4", str(rt3))
                                    child5.set("k5", str(rt2))
                                    child5.set("k6", str(rt1))
                                    #Write the new data to the same xml file or create a new with diferent name  
                                    tree.write(xmlfo)  
                                    #print(child5.attrib)
    
    def xml2exl(self,xmlf,exlf,exlfo,current_s,s_num):
        
        tree = ET.parse(xmlf)
        root = tree.getroot()
    
        #Open the workbook to read in
        x1_workbook = xlrd.open_workbook(exlf)
    
        #Obtain spreadsheet names of the current workbook
        sheet_names = x1_workbook.sheet_names()

        print('Sheet Names',sheet_names)
        
        #Call the first sheet of the book
        x1_sheet = x1_workbook.sheet_by_index(s_num)
       
        #Open the workbook to read in
        y1_workbook = load_workbook(exlf)
        wb = y1_workbook[current_s]
    
        #List variables to put the names of the bushing names contained in the xml and excel file
        xml_bushings = {}
    
        #Iterate over inline.xml file to obtain all the bushing names 
        for child in root.iter("NVHC_PROPERTY"):
            for child2 in child.iter("NAME"):
                #print(child2.text)
                for child3 in child.iter("PARAMETERS"):
                    #print(child3.tag)
                    for child4 in child3:
                        if child4.tag == "K_VALUES":
                            xml_bushings[child2.text] = child4.attrib
                            
        print(xml_bushings)
    
        for row_idx in range(0, x1_sheet.nrows):    # Iterate through rows
            for col_idx in range(0, x1_sheet.ncols):  # Iterate through columns
                cell_obj = x1_sheet.cell(row_idx, col_idx).value  # Get cell object by row, col 
                if cell_obj == "Bushing":
                    for row_idx2 in range(row_idx, x1_sheet.nrows): # Iterate through columns
                        for key, value in xml_bushings.items():
                            row_idx2 = row_idx2+1
                            v1 = xml_bushings[key]['k1']
                            v2 = xml_bushings[key]['k2'] 
                            v3 = xml_bushings[key]['k3'] 
                            v4 = xml_bushings[key]['k4'] 
                            v5 = xml_bushings[key]['k5'] 
                            v6 = xml_bushings[key]['k6']
                            wb.cell(row = row_idx2+1, column = col_idx+1, value = key)
                            print(wb.cell(row = row_idx2+1, column = col_idx+1).value)
                            for col_idx2 in range(col_idx+2,x1_sheet.ncols-1):
                                if(col_idx2 == 5):
                                    wb.cell(row = row_idx2+1, column = col_idx2, value = float(v1))
                                if(col_idx2 == 6):
                                    wb.cell(row = row_idx2+1, column = col_idx2, value = float(v2))
                                if(col_idx2 == 7):
                                    wb.cell(row = row_idx2+1, column = col_idx2, value = float(v3))
                                if(col_idx2 == 8):
                                    wb.cell(row = row_idx2+1, column = col_idx2, value = float(v4))
                                if(col_idx2 == 9):
                                    wb.cell(row = row_idx2+1, column = col_idx2, value = float(v5))
                                if(col_idx2 == 10):
                                    wb.cell(row = row_idx2+1, column = col_idx2, value = float(v6))
    
                        y1_workbook.save(exlfo)
                        print(exlfo)
                        break
                    messagebox.showinfo("Process", "Done")
                    
                    break